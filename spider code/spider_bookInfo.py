# -*- coding: utf-8 -*-
# @Author  : yocichen
# @Email   : yocichen@126.com
# @File    : doubanBooks.py
# @Software: PyCharm
# @Time    : 2019/11/9 11:38


'''
任务：
爬取书籍图片链接/
爬取内容简介/
爬取作者简介/
爬取标签（需要判断标签的类型，最好给一个筛选列表）- 改爬具体标签下的书籍src了
'''

import re
import openpyxl
import requests
from requests import RequestException
from bs4 import BeautifulSoup
import lxml
import time
import random

def get_one_page(url):
    '''
    Get the html of a page by requests module
    :param url: page url
    :return: html / None
    '''
    try:
        head = ['Mozilla/5.0', 'Chrome/78.0.3904.97', 'Safari/537.36']
        headers = {
            'user-agent':head[random.randint(0, 2)]
        }
        response = requests.get(url, headers=headers) #, proxies={'http':'171.15.65.195:9999'}
        if response.status_code == 200:
            return response.text
        return None
    except RequestException:
        return None

def get_request_res(pattern_text, html):
    '''
    Get the book info by re module
    :param pattern_text: re pattern
    :param html: page's html text
    :return: book's info
    '''
    pattern = re.compile(pattern_text, re.S)
    res = re.findall(pattern, html)
    if len(res) > 0:
        return res[0].split('<', 1)[0][1:]
    else:
        return 'NULL'

def get_bs_res(selector, html):
    '''
    Get the book info by bs4 module
    :param selector: info selector
    :param html: page's html text
    :return: book's info
    '''
    soup = BeautifulSoup(html, 'lxml')
    res = soup.select(selector)
    # if res is not None or len(res) is not 0:
    #     return res[0].string
    # else:
    #     return 'NULL'
    if res is None:
        return 'NULL'
    elif len(res) == 0:
        return 'NULL'
    else:
        return res[0].string

# Get other info by bs module
def get_bs_img_res(selector, html):
    soup = BeautifulSoup(html, 'lxml')
    res = soup.select(selector)
    if len(res) is not 0:
        return str(res[0])
    else:
        return 'NULL'

def parse_one_page(html):
    '''
    Parse the useful info of html by re module
    :param html: page's html text
    :return: all of book info(dict)
    '''
    book_info = {}
    book_name = get_bs_res('div > h1 > span', html)
    # print('Book-name', book_name)
    book_info['Book_name'] = book_name
    # info > a:nth-child(2)
    author = get_bs_res('div > span:nth-child(1) > a', html)
    if author is None:
        author = get_bs_res('#info > a:nth-child(2)', html)
    # print('Author', author)
    author = author.replace(" ", "")
    author = author.replace("\n", "")
    book_info['Author'] = author

    publisher = get_request_res(u'出版社:</span>(.*?)<br/>', html)
    # print('Publisher', publisher)
    book_info['publisher'] = publisher

    publish_time = get_request_res(u'出版年:</span>(.*?)<br/>', html)
    # print('Publish-time', publish_time)
    book_info['publish_time'] = publish_time

    ISBN = get_request_res(u'ISBN:</span>(.*?)<br/>', html)
    # print('ISBN', ISBN)
    book_info['ISBN'] = ISBN

    img_label = get_bs_img_res('#mainpic > a > img', html)
    pattern = re.compile('src="(.*?)"', re.S)
    img = re.findall(pattern, img_label)
    if len(img) is not 0:
        # print('img-src', img[0])
        book_info['img_src'] = img[0]
    else:
        # print('src not found')
        book_info['img_src'] = 'NULL'

    book_intro = get_bs_res('#link-report > div:nth-child(1) > div > p', html)
    # print('book introduction', book_intro)
    book_info['book_intro'] = book_intro

    author_intro = get_bs_res('#content > div > div.article > div.related_info > div:nth-child(4) > div > div > p', html)
    # print('author introduction', author_intro)
    book_info['author_intro'] = author_intro

    grade = get_bs_res('div > div.rating_self.clearfix > strong', html)
    if len(grade) == 1:
        # print('Score no mark')
        book_info['Score'] = 'NULL'
    else:
        # print('Score', grade[1:])
        book_info['Score'] = grade[1:]

    comment_num = get_bs_res('#interest_sectl > div > div.rating_self.clearfix > div > div.rating_sum > span > a > span', html)
    # print('commments', comment_num)
    book_info['commments'] = comment_num

    five_stars = get_bs_res('#interest_sectl > div > span:nth-child(5)', html)
    # print('5-stars', five_stars)
    book_info['5_stars'] = five_stars

    four_stars = get_bs_res('#interest_sectl > div > span:nth-child(9)', html)
    # print('4-stars', four_stars)
    book_info['4_stars'] = four_stars

    three_stars = get_bs_res('#interest_sectl > div > span:nth-child(13)', html)
    # print('3-stars', three_stars)
    book_info['3_stars'] = three_stars

    two_stars = get_bs_res('#interest_sectl > div > span:nth-child(17)', html)
    # print('2-stars', two_stars)
    book_info['2_stars'] = two_stars

    one_stars = get_bs_res('#interest_sectl > div > span:nth-child(21)', html)
    # print('1-stars', one_stars)
    book_info['1_stars'] = one_stars

    return book_info

def write_bookinfo_excel(book_info, file):
    '''
    Write book info into excel file
    :param book_info: a dict
    :param file: memory excel file
    :return: the num of successful item
    '''
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    sheet_row = ws.max_row
    sheet_col = ws.max_column
    i = sheet_row
    j = 1
    for key in book_info:
        ws.cell(i+1, j).value = book_info[key]
        j += 1
    done = ws.max_row - sheet_row
    wb.save(file)
    return done

def read_booksrc_get_info(src_file, info_file):
    '''
    Read the src file and access each src, parse html and write info into file
    :param src_file: src file
    :param info_file: memory file
    :return: the num of successful item
    '''
    wb = openpyxl.load_workbook(src_file)
    ws = wb.worksheets[0]
    row = ws.max_row
    done = 0
    for i in range(868, row+1):
        src = ws.cell(i, 1).value
        if src is None:
            continue
        html = get_one_page(str(src))
        book_info = parse_one_page(html)
        done += write_bookinfo_excel(book_info, info_file)
        if done % 10 == 0:
            print(done, 'done')
    return done

if __name__ == '__main__':
    # url = 'https://book.douban.com/subject/1770782/'
    # html = get_one_page(url)
    # # print(html)
    # book_info = parse_one_page(html)
    # print(book_info)
    # res = write_bookinfo_excel(book_info, 'novel_books_info.xlsx')
    # print(res, 'done')
    res = read_booksrc_get_info('masterpiece_books_src.xlsx', 'masterpiece_books_info.xlsx')
    print(res, 'done')