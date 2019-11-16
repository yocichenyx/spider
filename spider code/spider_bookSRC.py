# -*- coding: utf-8 -*-
# @Author  : yocichen
# @Email   : yocichen@126.com
# @File    : labelListBooks.py
# @Software: PyCharm
# @Time    : 2019/11/11 20:10

import re
import openpyxl
import requests
from requests import RequestException
from bs4 import BeautifulSoup
import lxml
import time
import random

src_list = []

def get_one_page(url):
    '''
    Get the html of a page by requests module
    :param url: page url
    :return: html / None
    '''
    try:
        head = ['Mozilla/5.0', 'Chrome/78.0.3904.97', 'Safari/537.36']
        headers = {
            'user-agent':head[random.randint(0, 2)]
        }
        response = requests.get(url, headers=headers, proxies={'http':'171.15.65.195:9999'})
        if response.status_code == 200:
            return response.text
        return None
    except RequestException:
        return None

def get_page_src(html, selector):
    '''
    Get book's src from label page
    :param html: book
    :param selector: src selector
    :return: src(list)
    '''
    # html = get_one_page(url)
    if html is not None:
        soup = BeautifulSoup(html, 'lxml')
        res = soup.select(selector)
        pattern = re.compile('href="(.*?)"', re.S)
        src = re.findall(pattern, str(res))
        return src
    else:
        return []

def write_excel_xlsx(items, file):
    '''
    Write the useful info into excel(*.xlsx file)
    :param items: book's info
    :param file: memory excel file
    :return: the num of successful item
    '''
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    sheet_row = ws.max_row
    item_num = len(items)
    # Write film's info
    for i in range(0, item_num):
        ws.cell(sheet_row+i+1, 1).value = items[i]
    # Save the work book as *.xlsx
    wb.save(file)
    return item_num

if __name__ == '__main__':
    total = 0
    for page_index in range(0, 50):
        # novel label src : https://book.douban.com/tag/%E5%B0%8F%E8%AF%B4?start=
        # program label src : https://book.douban.com/tag/%E7%BC%96%E7%A8%8B?start=
        # computer label src : https://book.douban.com/tag/%E8%AE%A1%E7%AE%97%E6%9C%BA?start=
        # masterpiece label src : https://book.douban.com/tag/%E5%90%8D%E8%91%97?start=
        url = 'https://book.douban.com/tag/%E5%90%8D%E8%91%97?start='+str(page_index*20)+'&type=T'
        one_loop_done = 0
        # only get html page once
        html = get_one_page(url)
        for book_index in range(1, 21):
            selector = '#subject_list > ul > li:nth-child('+str(book_index)+') > div.info > h2'
            src = get_page_src(html, selector)
            row = write_excel_xlsx(src, 'masterpiece_books_src.xlsx')
            one_loop_done += row
        total += one_loop_done
        print(one_loop_done, 'done')
    print('Total', total, 'done')