# -*- coding: utf-8 -*-
# @Author  : yocichen
# @Email   : yocichen@126.com
# @File    : MaoyanTop100.py
# @Software: PyCharm
# @Time    : 2019/11/6 9:52

import requests
from requests import RequestException
import re
import openpyxl

# Get page's html by requests module
def get_one_page(url):
    try:
        headers = {
            'user-agent':'Mozilla/5.0'
        }
        # use headers to avoid 403 Forbidden Error(reject spider)
        response = requests.get(url, headers=headers)
        if response.status_code == 200 :
            return response.text
        return None
    except RequestException:
        return None

# Get useful info from html of a page by re module
def parse_one_page(html):
    pattern = re.compile('<dd>.*?board-index.*?>(\d+)<.*?<a.*?title="(.*?)"'
                         +'.*?data-src="(.*?)".*?</a>.*?star">[\\s]*(.*?)[\\n][\\s]*</p>.*?'
                         +'releasetime">(.*?)</p>.*?integer">(.*?)</i>.*?'
                         +'fraction">(.*?)</i>.*?</dd>', re.S)
    items = re.findall(pattern, html)
    return items

# Main call
def main(url):
    page_html = get_one_page(url)
    parse_res = parse_one_page(page_html)
    return parse_res

# Write the useful info in excel(*.xlsx file)
def write_excel_xlsx(items):
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = len(items)
    cols = len(items[0])
    # First, write col's title.
    ws.cell(1, 1).value = '编号'
    ws.cell(1, 2).value = '片名'
    ws.cell(1, 3).value = '宣传图片'
    ws.cell(1, 4).value = '主演'
    ws.cell(1, 5).value = '上映时间'
    ws.cell(1, 6).value = '评分'
    # Write film's info
    for i in range(0, rows):
        for j in range(0, cols):
            # print(items[i-1][j-1])
            if j != 5:
                ws.cell(i+2, j+1).value = items[i][j]
            else:
                ws.cell(i+2, j+1).value = items[i][j]+items[i][j+1]
                break
    # Save the work book as *.xlsx
    wb.save('maoyan_top100.xlsx')

if __name__ == '__main__':
    res = []
    url = 'https://maoyan.com/board/4?'
    for i in range(0, 10):
        if i == 0:
            res = main(url)
        else:
            newUrl = url+'offset='+str(i*10)
            res.extend(main(newUrl))
    # test spider
    # for item in res:
    #     print(item)
    # test wirte to excel
    # res = [
    #     [1, 2, 3, 4, 9],
    #     [2, 3, 4, 5, 9],
    #     [4, 5, 6, 7, 9]
    # ]

    write_excel_xlsx(res)
